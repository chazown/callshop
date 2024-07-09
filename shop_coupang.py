from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 크롬 드라이버 옵션 설정
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

# 웹 사이트 열기
driver = webdriver.Chrome(options=chrome_options)
driver.get("https://www.coupang.com")
time.sleep(1)

# 쇼핑 메뉴 클릭하기
driver.find_element(By.CSS_SELECTOR, ".service_icon.type_shopping").click()
time.sleep(1)

# 새 창으로 전환하기
new_window = driver.window_handles[1]
driver.switch_to.window(new_window)

# 화면 최대화
driver.maximize_window()

# 팝업 창 닫기 시도
try:
    # WebDriverWait를 사용하여 팝업 닫기 버튼이 나타날 때까지 기다림
    close_popup_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "button._buttonArea_button_1jZae._buttonArea_close_34bcm"))
    )
    close_popup_button.click()
    print("팝업 창이 닫혔습니다.")
except Exception as e:
    print("팝업 창을 닫는 중 예외 발생:", e)

# 잠시 대기 (팝업 닫히는 동안)
time.sleep(1)

# 검색창에 '상품명' 입력하고 검색 실행
search = driver.find_element(
    By.CSS_SELECTOR, "input._searchInput_search_text_3CUDs")
search.click()
search.send_keys("오리지널슬링백")
search.send_keys(Keys.ENTER)

# 스크롤 전 높이
before_h = driver.execute_script("return window.scrollY")

# 무한 스크롤
while True:
    # 맨 아래로 스크롤 내린다.
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    # 스크롤 사이 페이지 로딩 시간
    time.sleep(1)
    # 스크롤 후 높이
    after_h = driver.execute_script("return window.scrollY")
    if after_h == before_h:
        break
    before_h = after_h
    
# 새 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "상품 정보"
ws.append(['상품명', '가격', '링크', '총 리뷰 수', '이미지'])

# 상품 정보 수집
items = driver.find_elements(By.CSS_SELECTOR, ".product_item__MDtDF")
for item in items:
    # 상품명 찾기
    names = item.find_elements(By.CSS_SELECTOR, ".product_title__Mmw2K")
    name = names[0].text if names else "이름 없음"
    # 가격 찾기
    prices = item.find_elements(By.CSS_SELECTOR, ".price_num__S2p_v")
    price = prices[0].text if prices else "가격 없음"
    # 링크 찾기
    links = item.find_elements(By.CSS_SELECTOR, ".product_title__Mmw2K > a")
    link = links[0].get_attribute('href') if links else "링크 없음"
    # 리뷰 수 찾기
    review_totals = item.find_elements(By.CSS_SELECTOR, ".product_num__fafe5")
    review_total = review_totals[0].text if review_totals else "리뷰 없음"
    # 이미지 찾기
    images = item.find_elements(
        By.CSS_SELECTOR, '.thumbnail_thumb__Bxb6Z > img')
    image = images[0].get_attribute('src') if images else "이미지 없음"
    ws.append([name, price, link, review_total, image])
col_num = 5
for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    img_url = row[col_num-1]  # 이미지 URL
    if img_url:
        try:
            response = requests.get(img_url)
            img = Image(BytesIO(response.content))
            img.width = 140
            img.height = 140
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = ""
            ws.add_image(img, f'E{row_num}')
            ws.row_dimensions[row_num].height = 140
        except Exception as e:
            print(f"{e}")
            
# 배경색 지정
color = PatternFill(start_color='808080',
                    end_color='808080', fill_type='solid')
for row in ws.iter_rows(min_row=1, max_row=1):  # 첫 번째 줄만 색칠
    for cell in row:
        cell.fill = color
        
# 열 너비 변경
ws.column_dimensions['A'].width = 30
ws.column_dimensions['E'].width = 20

# 수직 가운데 맞춤 설정
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center")
excel_file_path = "쿠팡정보.xlsx"
wb.save(excel_file_path)

#웹 종료
driver.quit()